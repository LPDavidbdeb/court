"""Read-only audit of legal evidence references.

The legal drafting layer uses two interchangeable forms for evidence:

* Markdown source names such as ``piece_pdf-1.md``;
* explicit Django identities such as ``PDFDocument id=1``.

This module normalizes both forms to a canonical ``(model, pk)`` identity,
while preserving every observed spelling and its provenance.  It never assigns
procedural exhibit numbers and never mutates source objects.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Iterator

from django.db.models import Model

from document_manager.models import Document
from email_manager.models import Email, EmailThread
from events.models import Event
from googlechat_manager.models import ChatMessage, ChatSequence
from pdf_manager.models import PDFDocument
from photos.models import Photo, PhotoDocument


MODEL_CLASSES: dict[str, type[Model]] = {
    "PDFDocument": PDFDocument,
    "Document": Document,
    "Email": Email,
    "EmailThread": EmailThread,
    "Event": Event,
    "Photo": Photo,
    "PhotoDocument": PhotoDocument,
    "ChatMessage": ChatMessage,
    "ChatSequence": ChatSequence,
}

DIRECT_MODEL_NAMES = {
    "pdfdocument": "PDFDocument",
    "pdfdocuments": "PDFDocument",
    "document": "Document",
    "documents": "Document",
    "email": "Email",
    "emails": "Email",
    "emailthread": "EmailThread",
    "emailthreads": "EmailThread",
    "event": "Event",
    "events": "Event",
    "photo": "Photo",
    "photos": "Photo",
    "photodocument": "PhotoDocument",
    "photodocuments": "PhotoDocument",
    "chatmessage": "ChatMessage",
    "chatmessages": "ChatMessage",
    "chatsequence": "ChatSequence",
    "chatsequences": "ChatSequence",
}

PIECE_PATTERNS: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"\bpiece_pdf-(?P<pk>\d+)(?:\.md)?\b", re.I), "PDFDocument"),
    (re.compile(r"\bpiece_document-(?P<pk>\d+)(?:\.md)?\b", re.I), "Document"),
    (re.compile(r"\bpiece_photodoc-(?P<pk>\d+)(?:\.md)?\b", re.I), "PhotoDocument"),
    (re.compile(r"\bpiece_photo-(?P<pk>\d+)(?:\.md)?\b", re.I), "Photo"),
    (re.compile(r"\bpiece_event-(?P<pk>\d+)(?:\.md)?\b", re.I), "Event"),
    (re.compile(r"\bpiece_chatsequence-(?P<pk>\d+)(?:\.md)?\b", re.I), "ChatSequence"),
)

THREAD_EMAIL_PATTERN = re.compile(
    r"\bpiece_thread-(?P<thread_pk>\d+)_email-(?P<pk>\d+)(?:\.md)?\b",
    re.I,
)
THREAD_PATTERN = re.compile(
    r"\bpiece_thread-(?P<pk>\d+)(?:_[\wÀ-ÿ-]+)?(?:\.md)?\b",
    re.I,
)
GENERIC_PIECE_PATTERN = re.compile(
    r"\bpiece_[\wÀ-ÿ….-]+(?:\.md)?\b",
    re.I,
)
DIRECT_REFERENCE_PATTERN = re.compile(
    r"(?<![\w])"
    r"(?P<model>PDFDocuments?|Documents?|EmailThreads?|Emails?|Events?|"
    r"PhotoDocuments?|Photos?|ChatMessages?|ChatSequences?)"
    r"\s+(?:(?:id|ids|pk|pks)\s*=?\s*)"
    r"(?P<ids>\d+(?:\s*(?:,|\bet\b|\bà\b|\bau\b|–|-|/)\s*\d+)*)",
    re.I,
)

ANALYTICAL_MARKERS = (
    "référence analytique exclue",
    "références analytiques exclues",
    "documents analytiques exclus",
    "analyses exclues",
    "autorité juridique exclue",
    "source de droit, non comme preuve",
)


@dataclass(frozen=True)
class ReferenceOccurrence:
    source_file: str
    source_format: str
    source_location: str
    section: str
    raw_reference: str
    reference_form: str
    model: str | None
    pk: int | None
    context_model: str | None = None
    context_pk: int | None = None
    classification: str = "probative_candidate"
    in_facts: bool = False
    in_pont_story: bool = False
    in_pont_table: bool = False
    in_axis_argument: bool = False
    in_relation_synthesis: bool = False

    @property
    def canonical_key(self) -> str:
        if self.model and self.pk is not None:
            return f"{self.model}:{self.pk}"
        return ""


def _normalize_space(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _classification(context: str) -> str:
    normalized = context.casefold()
    if any(marker in normalized for marker in ANALYTICAL_MARKERS):
        return "analytical_excluded"
    return "probative_candidate"


def _membership(context: str) -> dict[str, bool]:
    normalized = context.casefold()
    return {
        "in_facts": "faits" in normalized,
        "in_pont_story": "pont" in normalized and any(
            token in normalized for token in ("récit", "recit", "version procédurale")
        ),
        "in_pont_table": "pont" in normalized and any(
            token in normalized for token in ("tableau", "pièces nécessaires", "pieces nécessaires")
        ),
        "in_axis_argument": any(token in normalized for token in ("axe", "argument")),
        "in_relation_synthesis": any(
            token in normalized for token in ("relation", "synthèse", "synthese")
        ),
    }


def _make_occurrence(
    *,
    source_file: str,
    source_format: str,
    source_location: str,
    section: str,
    raw_reference: str,
    reference_form: str,
    model: str | None,
    pk: int | None,
    context: str,
    context_model: str | None = None,
    context_pk: int | None = None,
) -> ReferenceOccurrence:
    return ReferenceOccurrence(
        source_file=source_file,
        source_format=source_format,
        source_location=source_location,
        section=section,
        raw_reference=_normalize_space(raw_reference).strip("`"),
        reference_form=reference_form,
        model=model,
        pk=pk,
        context_model=context_model,
        context_pk=context_pk,
        classification=_classification(f"{section} {context}"),
        **_membership(context),
    )


def extract_references_from_text(
    text: object,
    *,
    source_file: str,
    source_format: str,
    source_location: str,
    section: str = "",
    context: str = "",
) -> list[ReferenceOccurrence]:
    """Extract canonical and unresolved evidence references from a text cell."""
    value = _normalize_space(text)
    if not value:
        return []
    # Markdown emphasis around identifiers (for example ``id= **174**``)
    # should not prevent resolution. Replacing with spaces preserves spans.
    matchable = value.replace("*", " ").replace("`", " ")

    occurrences: list[ReferenceOccurrence] = []
    occupied_spans: list[tuple[int, int]] = []

    def overlaps(span: tuple[int, int]) -> bool:
        return any(span[0] < end and span[1] > start for start, end in occupied_spans)

    for match in THREAD_EMAIL_PATTERN.finditer(matchable):
        occurrences.append(
            _make_occurrence(
                source_file=source_file,
                source_format=source_format,
                source_location=source_location,
                section=section,
                raw_reference=match.group(0),
                reference_form="piece_markdown",
                model="Email",
                pk=int(match.group("pk")),
                context_model="EmailThread",
                context_pk=int(match.group("thread_pk")),
                context=context,
            )
        )
        occupied_spans.append(match.span())

    for pattern, model_name in PIECE_PATTERNS:
        for match in pattern.finditer(matchable):
            if overlaps(match.span()):
                continue
            occurrences.append(
                _make_occurrence(
                    source_file=source_file,
                    source_format=source_format,
                    source_location=source_location,
                    section=section,
                    raw_reference=match.group(0),
                    reference_form="piece_markdown",
                    model=model_name,
                    pk=int(match.group("pk")),
                    context=context,
                )
            )
            occupied_spans.append(match.span())

    for match in THREAD_PATTERN.finditer(matchable):
        if overlaps(match.span()):
            continue
        occurrences.append(
            _make_occurrence(
                source_file=source_file,
                source_format=source_format,
                source_location=source_location,
                section=section,
                raw_reference=match.group(0),
                reference_form="piece_markdown",
                model="EmailThread",
                pk=int(match.group("pk")),
                context=context,
            )
        )
        occupied_spans.append(match.span())

    for match in DIRECT_REFERENCE_PATTERN.finditer(matchable):
        if overlaps(match.span()):
            continue
        model_name = DIRECT_MODEL_NAMES[match.group("model").casefold()]
        raw_ids = match.group("ids")
        # Ranges and slash-separated identifiers are deliberately retained as
        # unresolved collectives; expanding them would violate the source's
        # own caution about ambiguous or collective references.
        if re.search(r"\d\s*(?:à|au|–|-|/)\s*\d", raw_ids, re.I):
            occurrences.append(
                _make_occurrence(
                    source_file=source_file,
                    source_format=source_format,
                    source_location=source_location,
                    section=section,
                    raw_reference=match.group(0),
                    reference_form="collective_or_ambiguous",
                    model=None,
                    pk=None,
                    context=context,
                )
            )
        else:
            for pk_text in re.findall(r"\d+", raw_ids):
                occurrences.append(
                    _make_occurrence(
                        source_file=source_file,
                        source_format=source_format,
                        source_location=source_location,
                        section=section,
                        raw_reference=f"{match.group('model')} id={pk_text}",
                        reference_form="direct_tuple",
                        model=model_name,
                        pk=int(pk_text),
                        context=context,
                    )
                )
        occupied_spans.append(match.span())

    for match in GENERIC_PIECE_PATTERN.finditer(matchable):
        if overlaps(match.span()):
            continue
        occurrences.append(
            _make_occurrence(
                source_file=source_file,
                source_format=source_format,
                source_location=source_location,
                section=section,
                raw_reference=match.group(0),
                reference_form="unparsed_piece_reference",
                model=None,
                pk=None,
                context=context,
            )
        )

    # A cell whose entire value is an old procedural cote represents a real,
    # unresolved union member.  Cotes merely mentioned as aliases elsewhere
    # are context, not separate canonical candidates.
    cote_match = re.fullmatch(r"P-\d+", value, re.I)
    if cote_match:
        occurrences.append(
            _make_occurrence(
                source_file=source_file,
                source_format=source_format,
                source_location=source_location,
                section=section,
                raw_reference=value,
                reference_form="procedural_alias_only",
                model=None,
                pk=None,
                context=context,
            )
        )

    unique: dict[tuple, ReferenceOccurrence] = {}
    for occurrence in occurrences:
        key = (
            occurrence.raw_reference.casefold(),
            occurrence.model,
            occurrence.pk,
            occurrence.context_model,
            occurrence.context_pk,
        )
        unique[key] = occurrence
    return list(unique.values())


def _split_markdown_row(line: str) -> list[str]:
    stripped = line.strip().strip("|")
    return [cell.strip() for cell in re.split(r"(?<!\\)\|", stripped)]


def _iter_markdown(path: Path, root: Path) -> Iterator[ReferenceOccurrence]:
    section = ""
    reference_column: int | None = None
    in_primary_table = False
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        heading = re.match(r"^#{1,6}\s+(.+?)\s*$", line)
        if heading:
            section = heading.group(1)

        if line.lstrip().startswith("|") and "|" in line:
            cells = _split_markdown_row(line)
            normalized_headers = [cell.casefold() for cell in cells]
            if "référence interne exacte" in normalized_headers:
                reference_column = normalized_headers.index("référence interne exacte")
                in_primary_table = True
                continue
            if all(re.fullmatch(r":?-{2,}:?", cell) for cell in cells if cell):
                continue
            if in_primary_table and reference_column is not None:
                if reference_column < len(cells) and cells[reference_column]:
                    row_context = " | ".join(cells)
                    yield from extract_references_from_text(
                        cells[reference_column],
                        source_file=str(path.relative_to(root)),
                        source_format="markdown",
                        source_location=f"line:{line_number}",
                        section=section,
                        context=row_context,
                    )
                continue
        elif in_primary_table:
            in_primary_table = False
            reference_column = None

        yield from extract_references_from_text(
            line,
            source_file=str(path.relative_to(root)),
            source_format="markdown",
            source_location=f"line:{line_number}",
            section=section,
            context=line,
        )


def _iter_csv(path: Path, root: Path) -> Iterator[ReferenceOccurrence]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    reference_column: int | None = None
    for row_number, row in enumerate(rows, 1):
        normalized = [_normalize_space(cell).casefold() for cell in row]
        if "référence interne exacte" in normalized:
            reference_column = normalized.index("référence interne exacte")
            continue
        if reference_column is None or reference_column >= len(row):
            continue
        row_context = " | ".join(row)
        yield from extract_references_from_text(
            row[reference_column],
            source_file=str(path.relative_to(root)),
            source_format="csv",
            source_location=f"row:{row_number}",
            context=row_context,
        )


def _iter_xlsx(path: Path, root: Path) -> Iterator[ReferenceOccurrence]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment guard
        raise RuntimeError("openpyxl is required to audit XLSX annexes") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            reference_column: int | None = None
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                values = [_normalize_space(cell) for cell in row]
                normalized = [value.casefold() for value in values]
                if "référence interne exacte" in normalized:
                    reference_column = normalized.index("référence interne exacte")
                    continue
                if reference_column is None or reference_column >= len(values):
                    continue
                reference = values[reference_column]
                if not reference:
                    continue
                row_context = " | ".join(values)
                yield from extract_references_from_text(
                    reference,
                    source_file=str(path.relative_to(root)),
                    source_format="xlsx",
                    source_location=f"sheet:{worksheet.title};row:{row_number}",
                    section=worksheet.title,
                    context=row_context,
                )
    finally:
        workbook.close()


def collect_occurrences(input_dir: Path) -> list[ReferenceOccurrence]:
    root = input_dir.resolve()
    occurrences: list[ReferenceOccurrence] = []
    for path in sorted(root.iterdir()):
        if path.name.startswith("audit_") or not path.is_file():
            continue
        if path.suffix.casefold() == ".md":
            occurrences.extend(_iter_markdown(path, root))
        elif path.suffix.casefold() == ".csv":
            occurrences.extend(_iter_csv(path, root))
        elif path.suffix.casefold() == ".xlsx":
            occurrences.extend(_iter_xlsx(path, root))

    unique: dict[tuple, ReferenceOccurrence] = {}
    for occurrence in occurrences:
        key = (
            occurrence.source_file,
            occurrence.source_location,
            occurrence.raw_reference.casefold(),
            occurrence.model,
            occurrence.pk,
        )
        unique[key] = occurrence
    return sorted(
        unique.values(),
        key=lambda item: (
            item.source_file.casefold(),
            item.source_location,
            item.raw_reference.casefold(),
        ),
    )


def _piece_source_block(path: Path) -> str:
    """Return the source-identification block of a descriptive piece file."""
    lines = path.read_text(encoding="utf-8").splitlines()
    first_separator = next((index for index, line in enumerate(lines) if line.strip() == "---"), None)
    leading_block = lines[:first_separator] if first_separator is not None else lines[:40]
    leading_text = "\n".join(leading_block)
    if re.search(r"(?:\b(?:id|pk)\s*=\s*\d+|media/|pdf_documents/|photos/|evidence_files/)", leading_text, re.I):
        return leading_text

    for index, line in enumerate(lines):
        if re.match(r"^#{1,6}\s+Référence dans la base de données\s*$", line, re.I):
            block = []
            for candidate in lines[index + 1:]:
                if re.match(r"^#{1,6}\s+", candidate):
                    break
                block.append(candidate)
            return "\n".join(block)
    return leading_text


def _media_path_candidates(text: str) -> set[str]:
    candidates = set()
    for match in re.finditer(
        r"(?:media/)?(?P<path>(?:pdf_documents|photos|evidence_files)/[^`\s)>]+)",
        text,
        re.I,
    ):
        candidates.add(match.group("path").rstrip(".,;:"))
    return candidates


def _objects_for_media_path(media_path: str) -> list[tuple[str, int]]:
    matches: list[tuple[str, int]] = []
    for pk in PDFDocument.objects.filter(file=media_path).values_list("pk", flat=True):
        matches.append(("PDFDocument", pk))
    for pk in Document.objects.filter(file_source=media_path).values_list("pk", flat=True):
        matches.append(("Document", pk))
    for pk in Photo.objects.filter(file=media_path).values_list("pk", flat=True):
        matches.append(("Photo", pk))
    return matches


def resolve_descriptive_piece_occurrences(
    occurrences: list[ReferenceOccurrence],
    legal_dir: Path,
) -> list[ReferenceOccurrence]:
    """Resolve descriptive ``piece_*.md`` aliases from their source blocks.

    A descriptive file can resolve to one source or to several components.  In
    the latter case one canonical occurrence is emitted for every explicitly
    identified component, while the alias spelling is preserved.
    """
    cache: dict[str, list[tuple[str, int]]] = {}
    resolved: list[ReferenceOccurrence] = []
    for occurrence in occurrences:
        if occurrence.reference_form != "unparsed_piece_reference":
            resolved.append(occurrence)
            continue

        basename = occurrence.raw_reference
        if basename == "piece_….md":
            resolved.append(occurrence)
            continue
        if not basename.casefold().endswith(".md"):
            basename += ".md"
        piece_path = legal_dir / basename
        if not piece_path.is_file():
            resolved.append(occurrence)
            continue

        if basename not in cache:
            source_block = _piece_source_block(piece_path)
            source_occurrences = extract_references_from_text(
                source_block,
                source_file=basename,
                source_format="piece_source_block",
                source_location="source_block",
                context=source_block,
            )
            identities = {
                (item.model, item.pk)
                for item in source_occurrences
                if item.model and item.pk is not None
            }
            for media_path in _media_path_candidates(source_block):
                identities.update(_objects_for_media_path(media_path))
            cache[basename] = sorted(identities, key=lambda item: (item[0], item[1]))

        identities = cache[basename]
        if not identities:
            resolved.append(
                ReferenceOccurrence(
                    **{
                        **asdict(occurrence),
                        "reference_form": "aggregate_or_unmapped_piece_reference",
                    }
                )
            )
            continue
        form = "descriptive_piece_resolved" if len(identities) == 1 else "descriptive_piece_components"
        for model_name, pk in identities:
            resolved.append(
                ReferenceOccurrence(
                    **{
                        **asdict(occurrence),
                        "reference_form": form,
                        "model": model_name,
                        "pk": pk,
                    }
                )
            )
    return resolved


def _field_file_status(field_file) -> tuple[bool, str]:
    if not field_file or not getattr(field_file, "name", ""):
        return False, ""
    name = str(field_file.name)
    try:
        return bool(field_file.storage.exists(name)), name
    except Exception:
        try:
            return os.path.exists(field_file.path), name
        except Exception:
            return False, name


def _legacy_file_status(path_value: str | None) -> tuple[bool, str]:
    if not path_value:
        return False, ""
    return os.path.exists(path_value), os.path.basename(path_value)


def _aggregate_status(results: Iterable[tuple[bool, str]]) -> tuple[str, str]:
    checked = list(results)
    if not checked:
        return "no_attached_original", ""
    available = sum(1 for exists, _ in checked if exists)
    references = "; ".join(reference for _, reference in checked if reference)
    if available == len(checked):
        return "all_available", references
    if available:
        return "partially_available", references
    return "none_available", references


def object_original_status(obj: Model) -> tuple[str, str]:
    if isinstance(obj, PDFDocument):
        exists, reference = _field_file_status(obj.file)
        return ("available" if exists else "missing", reference)
    if isinstance(obj, Document):
        exists, reference = _field_file_status(obj.file_source)
        return ("available" if exists else "missing", reference)
    if isinstance(obj, Email):
        file_exists, file_reference = _field_file_status(obj.eml_file)
        if file_exists:
            return "available", file_reference
        legacy_exists, legacy_reference = _legacy_file_status(obj.eml_file_path)
        return ("available" if legacy_exists else "missing", legacy_reference or file_reference)
    if isinstance(obj, Photo):
        file_exists, file_reference = _field_file_status(obj.file)
        if file_exists:
            return "available", file_reference
        legacy_exists, legacy_reference = _legacy_file_status(obj.file_path)
        return ("available" if legacy_exists else "missing", legacy_reference or file_reference)
    if isinstance(obj, PhotoDocument):
        return _aggregate_status(object_original_status(photo)[0] == "available" and (True, str(photo.file.name or photo.file_path)) or (False, str(photo.file.name or photo.file_path)) for photo in obj.photos.all())
    if isinstance(obj, Event):
        photos = list(obj.linked_photos.all())
        if photos:
            return _aggregate_status(
                (
                    object_original_status(photo)[0] == "available",
                    object_original_status(photo)[1],
                )
                for photo in photos
            )
        if obj.linked_email_id:
            status, reference = object_original_status(obj.linked_email)
            return ("available" if status == "available" else "missing", reference)
        return "no_attached_original", ""
    if isinstance(obj, EmailThread):
        return _aggregate_status(
            (
                object_original_status(email)[0] == "available",
                object_original_status(email)[1],
            )
            for email in obj.emails.all()
        )
    if isinstance(obj, (ChatMessage, ChatSequence)):
        return "render_required", "database content"
    return "not_supported", ""


def _object_date(obj: Model) -> str:
    value = None
    if isinstance(obj, PDFDocument):
        value = obj.document_date
    elif isinstance(obj, Document):
        value = obj.document_original_date
    elif isinstance(obj, Email):
        value = obj.date_sent
    elif isinstance(obj, Event):
        value = obj.date
    elif isinstance(obj, Photo):
        value = obj.datetime_original or obj.datetime_utc or obj.date_folder
    elif isinstance(obj, PhotoDocument):
        value = obj.get_exhibit_date()
    elif isinstance(obj, ChatMessage):
        value = obj.timestamp
    elif isinstance(obj, ChatSequence):
        value = obj.start_date
    elif isinstance(obj, EmailThread):
        value = obj.emails.order_by("date_sent").values_list("date_sent", flat=True).first()
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return ""


def _object_title(obj: Model) -> str:
    if isinstance(obj, Email):
        return obj.subject or "[Sans sujet]"
    if isinstance(obj, EmailThread):
        return obj.subject or "[Sans sujet]"
    if hasattr(obj, "title"):
        return str(getattr(obj, "title") or "")
    if isinstance(obj, Event):
        return _normalize_space(obj.explanation)[:300]
    if isinstance(obj, Photo):
        return obj.file_name or os.path.basename(obj.file_path or "")
    if isinstance(obj, ChatMessage):
        return _normalize_space(obj.text_content)[:300]
    return _normalize_space(str(obj))[:300]


def audit_occurrences(occurrences: list[ReferenceOccurrence]) -> dict:
    grouped: dict[tuple[str, int], list[ReferenceOccurrence]] = defaultdict(list)
    unresolved: dict[str, list[ReferenceOccurrence]] = defaultdict(list)
    for occurrence in occurrences:
        if occurrence.classification == "analytical_excluded":
            continue
        if occurrence.model and occurrence.pk is not None:
            grouped[(occurrence.model, occurrence.pk)].append(occurrence)
        else:
            unresolved[occurrence.raw_reference.casefold()].append(occurrence)

    objects_by_key: dict[tuple[str, int], Model] = {}
    for model_name, model_class in MODEL_CLASSES.items():
        pks = sorted(pk for candidate_model, pk in grouped if candidate_model == model_name)
        if not pks:
            continue
        queryset = model_class.objects.filter(pk__in=pks)
        if model_class == Email:
            queryset = queryset.select_related("thread")
        elif model_class == Event:
            queryset = queryset.select_related("linked_email").prefetch_related("linked_photos")
        elif model_class == PhotoDocument:
            queryset = queryset.prefetch_related("photos")
        elif model_class == EmailThread:
            queryset = queryset.prefetch_related("emails")
        elif model_class == ChatSequence:
            queryset = queryset.prefetch_related("messages")
        for obj in queryset:
            objects_by_key[(model_name, obj.pk)] = obj

    canonical_rows: list[dict] = []
    alias_rows: list[dict] = []
    exception_rows: list[dict] = []
    for key in sorted(grouped, key=lambda item: (item[0], item[1])):
        model_name, pk = key
        group = grouped[key]
        obj = objects_by_key.get(key)
        raw_references = sorted({item.raw_reference for item in group}, key=str.casefold)
        source_files = sorted({item.source_file for item in group}, key=str.casefold)
        if obj:
            original_status, original_reference = object_original_status(obj)
            db_status = "found"
            object_date = _object_date(obj)
            object_title = _object_title(obj)
        else:
            original_status, original_reference = "not_checked", ""
            db_status = "missing"
            object_date = ""
            object_title = ""

        context_checks = []
        for item in group:
            if item.context_model == "EmailThread" and isinstance(obj, Email):
                context_checks.append(
                    "match" if obj.thread_id == item.context_pk else f"mismatch:{item.context_pk}"
                )
        context_status = ";".join(sorted(set(context_checks))) if context_checks else ""
        row = {
            "canonical_key": f"{model_name}:{pk}",
            "model": model_name,
            "pk": pk,
            "db_status": db_status,
            "original_status": original_status,
            "original_reference": original_reference,
            "object_date": object_date,
            "object_title": object_title,
            "context_status": context_status,
            "raw_references": " | ".join(raw_references),
            "source_files": " | ".join(source_files),
            "occurrence_count": len(group),
            "in_facts": any(item.in_facts for item in group),
            "in_pont_story": any(item.in_pont_story for item in group),
            "in_pont_table": any(item.in_pont_table for item in group),
            "in_axis_argument": any(item.in_axis_argument for item in group),
            "in_relation_synthesis": any(item.in_relation_synthesis for item in group),
        }
        canonical_rows.append(row)
        if len(raw_references) > 1:
            alias_rows.append(
                {
                    "canonical_key": row["canonical_key"],
                    "raw_references": row["raw_references"],
                    "count": len(raw_references),
                }
            )
        if db_status != "found" or original_status in {"missing", "none_available", "partially_available"} or context_status.startswith("mismatch"):
            exception_rows.append(
                {
                    "exception_type": (
                        "db_missing" if db_status != "found" else
                        "context_mismatch" if context_status.startswith("mismatch") else
                        "original_incomplete"
                    ),
                    "canonical_key": row["canonical_key"],
                    "raw_reference": row["raw_references"],
                    "source_files": row["source_files"],
                    "detail": context_status or original_status,
                }
            )

    unresolved_rows: list[dict] = []
    for normalized_reference, group in sorted(unresolved.items()):
        raw_references = sorted({item.raw_reference for item in group}, key=str.casefold)
        source_files = sorted({item.source_file for item in group}, key=str.casefold)
        forms = sorted({item.reference_form for item in group})
        row = {
            "raw_reference": " | ".join(raw_references),
            "reference_form": " | ".join(forms),
            "source_files": " | ".join(source_files),
            "occurrence_count": len(group),
        }
        unresolved_rows.append(row)
        exception_rows.append(
            {
                "exception_type": "unresolved_reference",
                "canonical_key": "",
                "raw_reference": row["raw_reference"],
                "source_files": row["source_files"],
                "detail": row["reference_form"],
            }
        )

    analytical_occurrences = [
        occurrence for occurrence in occurrences if occurrence.classification == "analytical_excluded"
    ]
    found_count = sum(row["db_status"] == "found" for row in canonical_rows)
    canonical_count = len(canonical_rows)
    model_counts: dict[str, int] = defaultdict(int)
    db_found_by_model: dict[str, int] = defaultdict(int)
    original_status_counts: dict[str, int] = defaultdict(int)
    for row in canonical_rows:
        model_counts[row["model"]] += 1
        if row["db_status"] == "found":
            db_found_by_model[row["model"]] += 1
        original_status_counts[row["original_status"]] += 1
    unresolved_by_form: dict[str, int] = defaultdict(int)
    for row in unresolved_rows:
        unresolved_by_form[row["reference_form"]] += 1
    summary = {
        "source_file_count": len({item.source_file for item in occurrences}),
        "occurrence_count": len(occurrences),
        "canonical_reference_count": canonical_count,
        "db_found_count": found_count,
        "db_missing_count": canonical_count - found_count,
        "db_coverage_percent": round((found_count / canonical_count * 100), 2) if canonical_count else 0,
        "unresolved_unique_count": len(unresolved_rows),
        "analytical_excluded_occurrence_count": len(analytical_occurrences),
        "alias_group_count": len(alias_rows),
        "exception_count": len(exception_rows),
        "model_counts": dict(sorted(model_counts.items())),
        "db_found_by_model": dict(sorted(db_found_by_model.items())),
        "original_status_counts": dict(sorted(original_status_counts.items())),
        "unresolved_by_form": dict(sorted(unresolved_by_form.items())),
        "in_facts_count": sum(row["in_facts"] for row in canonical_rows),
        "in_pont_story_count": sum(row["in_pont_story"] for row in canonical_rows),
        "in_pont_table_count": sum(row["in_pont_table"] for row in canonical_rows),
        "in_axis_argument_count": sum(row["in_axis_argument"] for row in canonical_rows),
        "in_relation_synthesis_count": sum(row["in_relation_synthesis"] for row in canonical_rows),
        "intersection_all_tracked_layers_count": sum(
            row["in_facts"]
            and row["in_pont_story"]
            and row["in_pont_table"]
            and row["in_axis_argument"]
            and row["in_relation_synthesis"]
            for row in canonical_rows
        ),
    }
    return {
        "summary": summary,
        "canonical": canonical_rows,
        "occurrences": [asdict(item) | {"canonical_key": item.canonical_key} for item in occurrences],
        "unresolved": unresolved_rows,
        "aliases": alias_rows,
        "exceptions": sorted(
            exception_rows,
            key=lambda item: (item["exception_type"], item["canonical_key"], item["raw_reference"]),
        ),
    }


def _write_csv(path: Path, rows: list[dict]) -> None:
    fieldnames = list(rows[0].keys()) if rows else ["status"]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows or [{"status": "aucune ligne"}])


def write_audit_reports(output_dir: Path, audit: dict) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "json": output_dir / "audit_preuve.json",
        "summary": output_dir / "resume_audit.json",
        "canonical": output_dir / "inventaire_canonique.csv",
        "occurrences": output_dir / "occurrences_references.csv",
        "exceptions": output_dir / "exceptions_references.csv",
        "aliases": output_dir / "alias_references.csv",
        "unresolved": output_dir / "references_non_resolues.csv",
    }
    paths["json"].write_text(
        json.dumps(audit, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    paths["summary"].write_text(
        json.dumps(audit["summary"], ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    _write_csv(paths["canonical"], audit["canonical"])
    _write_csv(paths["occurrences"], audit["occurrences"])
    _write_csv(paths["exceptions"], audit["exceptions"])
    _write_csv(paths["aliases"], audit["aliases"])
    _write_csv(paths["unresolved"], audit["unresolved"])
    return list(paths.values())


def reports_digest(paths: Iterable[Path]) -> str:
    digest = hashlib.sha256()
    for path in sorted(paths, key=lambda item: item.name):
        digest.update(path.name.encode("utf-8"))
        digest.update(path.read_bytes())
    return digest.hexdigest()
