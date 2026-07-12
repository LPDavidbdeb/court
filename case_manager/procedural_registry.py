"""Build a proposed procedural registry from the audited evidence universe.

This stage deliberately remains a proposal: it reads the provisional bordereau,
historic aliases and pleading placeholders, but does not edit the pleading or
write exhibit numbers to Django.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Iterator

from case_manager.evidence_audit import (
    MODEL_CLASSES,
    ReferenceOccurrence,
    _object_date,
    _object_title,
    extract_references_from_text,
    object_original_status,
    resolve_descriptive_piece_occurrences,
)


PLACEHOLDER_PATTERN = re.compile(r"P-\[([^\]]+)\]")
PROCEDURAL_ALIAS_PATTERN = re.compile(r"\bP-\d+\b", re.I)
RANGE_REFERENCE_PATTERN = re.compile(
    r"^(?P<model>ChatMessages?|Emails?|Events?|Photos?|PDFDocuments?|PhotoDocuments?)"
    r"\s+(?:id|ids|pk|pks)\s*=\s*(?P<ids>.+)$",
    re.I,
)


def _plain(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _normalized(value: str) -> str:
    decomposed = unicodedata.normalize("NFD", value)
    without_accents = "".join(char for char in decomposed if unicodedata.category(char) != "Mn")
    return re.sub(r"\s+", " ", without_accents.casefold()).strip()


def _markdown_cells(line: str) -> list[str]:
    return [cell.strip() for cell in re.split(r"(?<!\\)\|", line.strip().strip("|"))]


def parse_bordereau(path: Path) -> list[dict]:
    rows: list[dict] = []
    in_table = False
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        if not line.lstrip().startswith("|"):
            if in_table:
                in_table = False
            continue
        cells = _markdown_cells(line)
        if cells and cells[0].casefold() == "cote prov.":
            in_table = True
            continue
        if not in_table or len(cells) < 5 or re.fullmatch(r":?-{2,}:?", cells[0]):
            continue
        if not re.fullmatch(r"P-\d+", cells[0], re.I):
            continue
        rows.append(
            {
                "proposed_cote": cells[0].upper(),
                "date_display": cells[1],
                "description": re.sub(r"[*`]", "", cells[2]),
                "support_reference": re.sub(r"[*⚠️()]", "", cells[3]).strip(),
                "source_base": cells[4],
                "bordereau_line": line_number,
            }
        )
    return rows


def resolve_reference_text(text: str, legal_dir: Path, source: str) -> list[ReferenceOccurrence]:
    occurrences = extract_references_from_text(
        text,
        source_file=source,
        source_format="procedural_registry",
        source_location="reference",
        context=text,
    )
    slash_emails = re.findall(r"(?<!thread-)email-(\d+(?:/\d+)+)", text, re.I)
    for email_group in slash_emails:
        occurrences.extend(
            extract_references_from_text(
                "Emails id=" + ", ".join(email_group.split("/")),
                source_file=source,
                source_format="procedural_registry",
                source_location="slash_email_components",
                context=text,
            )
        )
    return resolve_descriptive_piece_occurrences(occurrences, legal_dir)


def build_proposed_registry(
    bordereau_rows: list[dict],
    audit: dict,
    legal_dir: Path,
) -> tuple[list[dict], list[dict], list[dict]]:
    audit_by_key = {row["canonical_key"]: row for row in audit["canonical"]}
    missing_audit_keys: dict[str, dict] = {}
    registry_rows: list[dict] = []
    component_rows: list[dict] = []
    cotes_by_key: dict[str, set[str]] = defaultdict(set)

    for bordereau_row in bordereau_rows:
        occurrences = resolve_reference_text(
            bordereau_row["support_reference"],
            legal_dir,
            "bordereau_pieces.md",
        )
        canonical_keys = sorted(
            {
                occurrence.canonical_key
                for occurrence in occurrences
                if occurrence.canonical_key
            }
        )
        unresolved = sorted(
            {
                occurrence.raw_reference
                for occurrence in occurrences
                if not occurrence.canonical_key
            },
            key=str.casefold,
        )
        status = "resolved" if canonical_keys and not unresolved else "partial" if canonical_keys else "unresolved"
        registry_rows.append(
            {
                **bordereau_row,
                "registry_status": status,
                "component_count": len(canonical_keys),
                "canonical_keys": " | ".join(canonical_keys),
                "unresolved_references": " | ".join(unresolved),
                "is_liasse": len(canonical_keys) > 1,
            }
        )
        for order, canonical_key in enumerate(canonical_keys, 1):
            audit_row = audit_by_key.get(canonical_key, {})
            if not audit_row:
                if canonical_key not in missing_audit_keys:
                    model_name, pk_text = canonical_key.split(":", 1)
                    model_class = MODEL_CLASSES[model_name]
                    obj = model_class.objects.filter(pk=int(pk_text)).first()
                    if obj:
                        original_status, original_reference = object_original_status(obj)
                        missing_audit_keys[canonical_key] = {
                            "model": model_name,
                            "pk": int(pk_text),
                            "db_status": "found_outside_union_audit",
                            "original_status": original_status,
                            "original_reference": original_reference,
                            "object_date": _object_date(obj),
                            "object_title": _object_title(obj),
                        }
                    else:
                        missing_audit_keys[canonical_key] = {
                            "model": model_name,
                            "pk": int(pk_text),
                            "db_status": "missing",
                            "original_status": "not_checked",
                            "object_date": "",
                            "object_title": "",
                        }
                audit_row = missing_audit_keys[canonical_key]
            cotes_by_key[canonical_key].add(bordereau_row["proposed_cote"])
            component_rows.append(
                {
                    "proposed_cote": bordereau_row["proposed_cote"],
                    "component_order": order,
                    "canonical_key": canonical_key,
                    "model": audit_row.get("model", canonical_key.split(":", 1)[0]),
                    "pk": audit_row.get("pk", canonical_key.split(":", 1)[1]),
                    "db_status": audit_row.get("db_status", "not_in_audit"),
                    "original_status": audit_row.get("original_status", "not_in_audit"),
                    "object_date": audit_row.get("object_date", ""),
                    "object_title": audit_row.get("object_title", ""),
                    "support_reference": bordereau_row["support_reference"],
                }
            )

    conflict_rows = [
        {
            "canonical_key": canonical_key,
            "proposed_cotes": " | ".join(sorted(cotes, key=_cote_sort_key)),
            "cote_count": len(cotes),
            "status": "same_tuple_multiple_proposed_cotes",
        }
        for canonical_key, cotes in sorted(cotes_by_key.items())
        if len(cotes) > 1
    ]
    return registry_rows, component_rows, conflict_rows


def _cote_sort_key(value: str) -> tuple[int, str]:
    match = re.search(r"\d+", value)
    return (int(match.group()) if match else 10**9, value)


def _iter_markdown_union_rows(path: Path) -> Iterator[dict]:
    header: list[str] | None = None
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        if not line.lstrip().startswith("|"):
            header = None
            continue
        cells = _markdown_cells(line)
        normalized = [_normalized(cell) for cell in cells]
        if "reference interne exacte" in normalized and "cote ou alias mentionne" in normalized:
            header = normalized
            continue
        if header is None or all(re.fullmatch(r":?-{2,}:?", cell) for cell in cells if cell):
            continue
        if len(cells) < len(header):
            cells.extend([""] * (len(header) - len(cells)))
        yield {
            "source_file": path.name,
            "source_location": f"line:{line_number}",
            "exact_reference": cells[header.index("reference interne exacte")],
            "alias_text": cells[header.index("cote ou alias mentionne")],
        }


def _iter_csv_union_rows(path: Path) -> Iterator[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        header: list[str] | None = None
        for row_number, cells in enumerate(reader, 1):
            normalized = [_normalized(cell) for cell in cells]
            if "reference interne exacte" in normalized and "cote ou alias mentionne" in normalized:
                header = normalized
                continue
            if header is None:
                continue
            if len(cells) <= max(header.index("reference interne exacte"), header.index("cote ou alias mentionne")):
                continue
            yield {
                "source_file": path.name,
                "source_location": f"row:{row_number}",
                "exact_reference": cells[header.index("reference interne exacte")],
                "alias_text": cells[header.index("cote ou alias mentionne")],
            }


def _iter_xlsx_union_rows(path: Path) -> Iterator[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            header: list[str] | None = None
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                cells = [_plain(cell) for cell in row]
                normalized = [_normalized(cell) for cell in cells]
                if "reference interne exacte" in normalized and "cote ou alias mentionne" in normalized:
                    header = normalized
                    continue
                if header is None:
                    continue
                exact_index = header.index("reference interne exacte")
                alias_index = header.index("cote ou alias mentionne")
                if max(exact_index, alias_index) >= len(cells) or not cells[exact_index]:
                    continue
                yield {
                    "source_file": path.name,
                    "source_location": f"sheet:{worksheet.title};row:{row_number}",
                    "exact_reference": cells[exact_index],
                    "alias_text": cells[alias_index],
                }
    finally:
        workbook.close()


def collect_historic_alias_mappings(input_dir: Path, legal_dir: Path) -> list[dict]:
    candidates: dict[str, dict[str, set[str]]] = defaultdict(lambda: {"keys": set(), "sources": set()})
    for path in sorted(input_dir.iterdir()):
        if not path.is_file():
            continue
        if path.suffix.casefold() == ".md":
            rows = _iter_markdown_union_rows(path)
        elif path.suffix.casefold() == ".csv":
            rows = _iter_csv_union_rows(path)
        elif path.suffix.casefold() == ".xlsx":
            rows = _iter_xlsx_union_rows(path)
        else:
            continue
        for row in rows:
            aliases = sorted(set(PROCEDURAL_ALIAS_PATTERN.findall(row["alias_text"])), key=_cote_sort_key)
            if not aliases:
                continue
            occurrences = resolve_reference_text(row["exact_reference"], legal_dir, row["source_file"])
            keys = {occurrence.canonical_key for occurrence in occurrences if occurrence.canonical_key}
            for alias in aliases:
                normalized_alias = alias.upper()
                candidates[normalized_alias]["keys"].update(keys)
                candidates[normalized_alias]["sources"].add(f"{row['source_file']}:{row['source_location']}")

    rows = []
    for alias, data in sorted(candidates.items(), key=lambda item: _cote_sort_key(item[0])):
        keys = sorted(data["keys"])
        rows.append(
            {
                "historic_alias": alias,
                "mapping_status": "resolved" if len(keys) == 1 else "ambiguous" if len(keys) > 1 else "unresolved",
                "candidate_count": len(keys),
                "canonical_candidates": " | ".join(keys),
                "source_locations": " | ".join(sorted(data["sources"])),
            }
        )
    return rows


def expand_collective_references(unresolved_rows: list[dict], audit: dict) -> list[dict]:
    audit_by_key = {row["canonical_key"]: row for row in audit["canonical"]}
    model_names = {
        "chatmessage": "ChatMessage", "chatmessages": "ChatMessage",
        "email": "Email", "emails": "Email",
        "event": "Event", "events": "Event",
        "photo": "Photo", "photos": "Photo",
        "pdfdocument": "PDFDocument", "pdfdocuments": "PDFDocument",
        "photodocument": "PhotoDocument", "photodocuments": "PhotoDocument",
    }
    rows: list[dict] = []
    for unresolved in unresolved_rows:
        if "collective_or_ambiguous" not in unresolved["reference_form"]:
            continue
        raw = unresolved["raw_reference"]
        match = RANGE_REFERENCE_PATTERN.match(raw)
        if not match:
            continue
        model = model_names[match.group("model").casefold()]
        expression = match.group("ids")
        if "/" in expression:
            rows.append({
                "collective_reference": raw, "model": model, "candidate_pk": "",
                "canonical_key": "", "db_status": "not_expanded", "expansion_status": "ambiguous_slash",
                "source_files": unresolved["source_files"],
            })
            continue
        ranges = list(re.finditer(r"(\d+)\s*(?:à|au|–|-)\s*(\d+)", expression, re.I))
        explicit_ids = {int(value) for value in re.findall(r"\d+", expression)}
        expanded_ids: set[int] = set()
        oversized = False
        for range_match in ranges:
            start, end = int(range_match.group(1)), int(range_match.group(2))
            if end < start or end - start > 50:
                oversized = True
                continue
            expanded_ids.update(range(start, end + 1))
            explicit_ids.discard(start)
            explicit_ids.discard(end)
        if oversized:
            rows.append({
                "collective_reference": raw, "model": model, "candidate_pk": "",
                "canonical_key": "", "db_status": "not_expanded", "expansion_status": "range_too_large_or_invalid",
                "source_files": unresolved["source_files"],
            })
            continue
        expanded_ids.update(explicit_ids)
        for pk in sorted(expanded_ids):
            key = f"{model}:{pk}"
            rows.append({
                "collective_reference": raw,
                "model": model,
                "candidate_pk": pk,
                "canonical_key": key,
                "db_status": audit_by_key.get(key, {}).get("db_status", "not_in_union_audit"),
                "expansion_status": "proposed_requires_review",
                "source_files": unresolved["source_files"],
            })
    return rows


STRONG_PLACEHOLDER_MAPPINGS = {
    "requete du 19 novembre 2015": ("P-19", "description_explicit"),
    "requete-2015": ("P-19", "description_explicit"),
    "courriel du 11 juin 2013": ("P-2", "description_explicit"),
    "lettre du 27 avril 2015": ("P-9", "description_explicit"),
    "lettre-27-avril": ("P-9", "description_explicit"),
    "messages textes du 7 avril 2015": ("P-8", "description_explicit"),
    "textos": ("P-8", "description_explicit"),
    "courriel et messages": ("P-8", "description_explicit"),
    "echange du 16 septembre 2016": ("P-22", "description_explicit"),
    "projet-13-aout": ("P-16", "description_explicit"),
    "reponse-2-septembre": ("P-17", "description_explicit"),
    "reponse-3-septembre": ("P-18", "description_explicit"),
    "jugement-2016": ("P-21", "description_explicit"),
    "p-1": ("P-2", "historic_alias_to_new_cote"),
    "ebauche du 4 mars 2015": ("P-7", "piece_pdf-2_body_date"),
    "lettre du 20 avril 2015": ("P-7", "piece_pdf-2_transmission_date"),
    "offre-20-avril": ("P-7", "piece_pdf-2_transmission_date"),
    "preuve du depart du 23 fevrier 2015": ("P-9", "piece_pdf-3_explicit_depart_date"),
    "da-2019": ("P-42", "document_3_sworn_declaration"),
}


def extract_pleading_placeholders(path: Path) -> list[dict]:
    rows: list[dict] = []
    current_paragraph = ""
    current_heading = ""
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        heading = re.match(r"^#{1,6}\s+(.+)$", line)
        if heading:
            current_heading = heading.group(1)
        paragraph = re.match(r"^(\d+)\.\s+", line)
        if paragraph:
            current_paragraph = paragraph.group(1)
        matches = list(PLACEHOLDER_PATTERN.finditer(line))
        for occurrence_index, match in enumerate(matches, 1):
            label = match.group(1)
            normalized = _normalized(label)
            proposed_cote = ""
            mapping_status = "unresolved"
            mapping_basis = ""
            if normalized in STRONG_PLACEHOLDER_MAPPINGS:
                proposed_cote, mapping_basis = STRONG_PLACEHOLDER_MAPPINGS[normalized]
                mapping_status = "proposed_strong"
            elif normalized in {"●"} or normalized.isdigit():
                mapping_status = "context_required"
                mapping_basis = "generic_or_section_local_placeholder"
            rows.append(
                {
                    "line": line_number,
                    "paragraph": current_paragraph,
                    "heading": current_heading,
                    "placeholder": f"P-[{label}]",
                    "occurrence_in_line": occurrence_index,
                    "mapping_status": mapping_status,
                    "proposed_cote": proposed_cote,
                    "mapping_basis": mapping_basis,
                    "context": _plain(line),
                }
            )
    return rows


def build_registry_report(
    *,
    audit: dict,
    bordereau_rows: list[dict],
    registry_rows: list[dict],
    component_rows: list[dict],
    conflict_rows: list[dict],
    alias_rows: list[dict],
    range_rows: list[dict],
    placeholder_rows: list[dict],
) -> dict:
    summary = {
        "proposed_cote_count": len(bordereau_rows),
        "resolved_cote_count": sum(row["registry_status"] == "resolved" for row in registry_rows),
        "partial_cote_count": sum(row["registry_status"] == "partial" for row in registry_rows),
        "unresolved_cote_count": sum(row["registry_status"] == "unresolved" for row in registry_rows),
        "liasse_cote_count": sum(bool(row["is_liasse"]) for row in registry_rows),
        "component_row_count": len(component_rows),
        "tuple_multiple_cote_conflict_count": len(conflict_rows),
        "historic_alias_count": len(alias_rows),
        "historic_alias_resolved_count": sum(row["mapping_status"] == "resolved" for row in alias_rows),
        "historic_alias_ambiguous_count": sum(row["mapping_status"] == "ambiguous" for row in alias_rows),
        "historic_alias_unresolved_count": sum(row["mapping_status"] == "unresolved" for row in alias_rows),
        "collective_expansion_row_count": len(range_rows),
        "collective_candidate_found_count": sum(row["db_status"] == "found" for row in range_rows),
        "collective_candidate_outside_union_count": sum(row["db_status"] == "not_in_union_audit" for row in range_rows),
        "collective_not_expanded_count": sum(row["db_status"] == "not_expanded" for row in range_rows),
        "pleading_placeholder_count": len(placeholder_rows),
        "pleading_placeholder_strong_mapping_count": sum(row["mapping_status"] == "proposed_strong" for row in placeholder_rows),
        "pleading_placeholder_context_required_count": sum(row["mapping_status"] == "context_required" for row in placeholder_rows),
        "pleading_placeholder_missing_from_bordereau_count": sum(row["mapping_status"] == "missing_from_bordereau" for row in placeholder_rows),
        "pleading_placeholder_unresolved_count": sum(row["mapping_status"] == "unresolved" for row in placeholder_rows),
        "component_db_missing_count": sum(row["db_status"] == "missing" for row in component_rows),
        "component_found_outside_union_count": sum(row["db_status"] == "found_outside_union_audit" for row in component_rows),
    }
    return {
        "summary": summary,
        "registry": registry_rows,
        "components": component_rows,
        "conflicts": conflict_rows,
        "historic_aliases": alias_rows,
        "collective_expansions": range_rows,
        "pleading_placeholders": placeholder_rows,
        "audit_summary": audit["summary"],
    }


def _write_csv(path: Path, rows: list[dict]) -> None:
    fieldnames = list(rows[0]) if rows else ["status"]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows or [{"status": "aucune ligne"}])


def write_registry_reports(output_dir: Path, report: dict) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "json": output_dir / "registre_procedural.json",
        "summary": output_dir / "resume_registre.json",
        "registry": output_dir / "registre_procedural_propose.csv",
        "components": output_dir / "composantes_cotes.csv",
        "conflicts": output_dir / "conflits_cotes.csv",
        "aliases": output_dir / "aliases_historiques.csv",
        "ranges": output_dir / "plages_collectives.csv",
        "placeholders": output_dir / "placeholders_demande.csv",
    }
    paths["json"].write_text(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    paths["summary"].write_text(json.dumps(report["summary"], ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    _write_csv(paths["registry"], report["registry"])
    _write_csv(paths["components"], report["components"])
    _write_csv(paths["conflicts"], report["conflicts"])
    _write_csv(paths["aliases"], report["historic_aliases"])
    _write_csv(paths["ranges"], report["collective_expansions"])
    _write_csv(paths["placeholders"], report["pleading_placeholders"])
    return list(paths.values())


def reports_digest(paths: list[Path]) -> str:
    digest = hashlib.sha256()
    for path in sorted(paths, key=lambda item: item.name):
        digest.update(path.name.encode("utf-8"))
        digest.update(path.read_bytes())
    return digest.hexdigest()
